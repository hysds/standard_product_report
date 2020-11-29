#!/usr/bin/env python

'''
Generates the Standard Product Ops Report
'''
from __future__ import print_function
from builtins import range
import re
import os
import json
import shutil
import urllib3
import hashlib
import datetime
import requests
from openpyxl import Workbook
import dateutil.parser
from hysds.celery import app

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

VERSION = 'v2.0'
PRODUCT_NAME = 'AOI_Ops_Report-{}-TN{}-{}-{}'
IDX_DCT = {'audit_trail': 'grq_*_runconfig-acqlist-audit_trail', 'ifg':'grq_*_s1-gunw',
           'acq-list':'grq_*_runconfig-acq-list', 'runconfig-topsapp': 'grq_*_runconfig-topsapp',
           'ifg-blacklist':'grq_*_blacklist', 'slc': 'grq_*_s1-iw_slc-local', 
           'acq': 'grq_*_acquisition-s1-iw_slc', 'aoi_track': 'grq_*_s1-gunw-aoi_track'}

def main():
    '''
    Queries for relevant products & builds the report by track.
    '''
    ctx = load_context()
    aoi_id = ctx.get('aoi_id', False)
    aoi_index = ctx.get('aoi_index', False)
    if aoi_id is False or aoi_index is False:
        raise Exception('invalid inputs of aoi_id: {}, aoi_index: {}'.format(aoi_id, aoi_index))
    aoi = get_aoi(aoi_id, aoi_index)
    track_acq_lists = sort_by_track(get_objects('acq-list', aoi))
    for track in list(track_acq_lists.keys()):
        print('For track: {}'.format(track))
        acqs = get_objects('acq', aoi, track)
        slcs = get_objects('slc', aoi, track)
        audit_trail = get_objects('audit_trail', aoi, track)
        if len(audit_trail) < 1:
            print('no audit trail products found for track {}'.format(track))
            continue
        allowed_hashes = list(set(store_by_hash(audit_trail).keys())) #allow only hashes foud in audit-trail
        acq_lists = filter_hashes(get_objects('acq-list', aoi, track), allowed_hashes)
        ifg_cfgs = filter_hashes(get_objects('runconfig-topsapp', aoi, track), allowed_hashes)
        ifgs = filter_hashes(get_objects('ifg', aoi, track), allowed_hashes)
        aoi_tracks = get_objects('aoi_track', aoi, track)
        now = datetime.datetime.now().strftime('%Y%m%dT%H%M')
        product_id = PRODUCT_NAME.format(aoi_id, track, now, VERSION)
        generate(product_id, aoi, track, acqs, slcs, acq_lists, ifg_cfgs, ifgs, audit_trail, aoi_tracks)
        print('generated {} for track: {}'.format(product_id, track))

def generate(product_id, aoi, track, acqs, slcs, acq_lists, ifg_cfgs, ifgs, audit_trail, aoi_tracks):
    '''generates an enumeration comparison report for the given aoi & track'''
    # unique tracks based on acquisition list
    if os.path.exists(product_id):
        shutil.rmtree(product_id)
    os.mkdir(product_id)
    filename = '{}.xlsx'.format(product_id)
    output_path = os.path.join(product_id, filename)
    acq_dct = store_by_id(acqs)
    acq_map_dct = store_by_slc_id(acqs)
    slc_dct = store_by_id(slcs)
    acq_list_dct = store_by_hash(acq_lists) # converts dict where key is hash of master/slave slc ids
    ifg_cfg_dct = store_by_hash(ifg_cfgs) # converts dict where key is hash of master/slave slc ids
    ifg_dct = store_by_hash(ifgs) # converts dict where key is hash of master/slave slc ids
    aoi_track_dct = store_by_gunw(aoi_tracks)
    #create workbook
    wb = Workbook()
    write_current_status(wb, acq_list_dct, ifg_cfg_dct, ifg_dct, slc_dct, acq_map_dct, aoi_track_dct)
    write_slcs(wb, slc_dct)
    write_missing_slcs(wb, slc_dct, acq_lists)
    write_acqs(wb, acq_dct)
    write_acq_lists(wb, acq_list_dct)
    write_ifg_cfgs(wb, ifg_cfg_dct)
    write_ifgs(wb, ifg_dct)
    #save output 
    wb.save(output_path)
    gen_product_met(aoi, product_id, track)

def write_current_status(wb, acq_list_dict, ifg_cfg_dct, ifg_dct, slc_dct, acq_map_dct, aoi_track_dct):
    '''generate the sheet for enumerated products'''
    ws = wb.active
    ws.title = 'Current Product Status'
    title = ['date pair', 'acquisition-list', 'runconfig-topsapp', 'ifg', 'hash', 'missing_slc_ids', 'missing_acq_ids', 'aoi_track_id']
    ws.append(title)
    for id_hash in sort_into_hash_list(acq_list_dict):
        acq_list = acq_list_dict.get(id_hash, {})
        ifg_cfg = ifg_cfg_dct.get(id_hash, {})
        ifg_cfg_id = ifg_cfg.get('_id', 'MISSING')
        ifg = ifg_dct.get(id_hash, {})
        date_pair = gen_date_pair(acq_list)
        acq_list_id = acq_list.get('_id', 'MISSING')
        ifg_id = ifg.get('_id', 'MISSING')
        aoi_track_id = aoi_track_dct.get(ifg_id, 'MISSING')
        missing_slcs = []
        missing_acqs = []
        acq_list_slcs = acq_list.get('_source').get('metadata').get('master_scenes') + acq_list.get('_source').get('metadata').get('slave_scenes')
        for slc_id in acq_list_slcs:
            if not slc_dct.get(slc_id, False):
                missing_slcs.append(slc_id)
                missing_acq = acq_map_dct.get(slc_id, False)
                if missing_acq:
                    missing_acq_id = missing_acq.get('_id')
                    missing_acqs.append(missing_acq_id)
        missing_slc_str = ', '.join(missing_slcs)
        missing_acq_str = ', '.join(missing_acqs) 
        ws.append([date_pair, acq_list_id, ifg_cfg_id, ifg_id, id_hash, missing_slc_str, missing_acq_str, aoi_track_id])

def write_slcs(wb, slc_dct):
    '''generates the sheet for slcs'''
    ws = wb.create_sheet('SLCs')
    ws.append(['slc_id'])
    for slc_id in list(slc_dct.keys()):
        ws.append([slc_id])

def write_missing_slcs(wb, slc_dct, acq_lists):
    '''generates the sheet for missing slcs'''
    ws = wb.create_sheet('Missing SLCs')
    ws.append(['slc_id'])
    missing = []
    for acq_list in acq_lists:
        master_scenes = acq_list.get('_source', {}).get('metadata', {}).get('master_scenes', [])
        slave_scenes = acq_list.get('_source', {}).get('metadata', {}).get('slave_scenes', [])
        all_scenes = master_scenes + slave_scenes
        for slc_id in all_scenes:
            if slc_dct.get(slc_id, False) is False:
                missing.append(slc_id)
    missing = list(set(missing))
    for slc_id in missing:
        ws.append([slc_id])

def write_acqs(wb, acq_dct):
    '''generates the sheet for acquisitions'''
    ws = wb.create_sheet('Acquisitions')
    ws.append(['acq_id', 'slc_id', 'ipf'])
    for acq_id in list(acq_dct.keys()):
        acq = acq_dct.get(acq_id, {})
        slc_id = acq.get('_source', {}).get('metadata', {}).get('title', 'MISSING')
        ipf = acq.get('_source', {}).get('metadata', {}).get('processing_version', 'MISSING')
        ws.append([acq_id, slc_id, ipf])

def write_acq_lists(wb, acq_list_dct):
    '''generates the sheet for acquisition lists'''
    ws = wb.create_sheet('Acquisition-Lists')
    ws.append(['acq_list_id', 'hash'])
    for hash_id in list(acq_list_dct.keys()):
        acq_list = acq_list_dct.get(hash_id, {})
        acq_id = acq_list.get('_source', {}).get('id', 'MISSING')
        ws.append([acq_id, hash_id])

def write_ifg_cfgs(wb, ifg_cfg_dct):
    '''generates the sheet for ifg cfgs'''
    ws = wb.create_sheet('IFG-Configs')
    ws.append(['ifg_cfg_id', 'hash'])
    for hash_id in list(ifg_cfg_dct.keys()):
        ifg_cfg = ifg_cfg_dct.get(hash_id, {})
        ifg_cfg_id = ifg_cfg.get('_source', {}).get('id', 'MISSING')
        ws.append([ifg_cfg_id, hash_id])

def write_ifgs(wb, ifg_dct):
    '''generates the sheet for ifgs'''
    ws = wb.create_sheet('IFGs')
    ws.append(['ifg_cfg_id', 'hash'])
    for hash_id in list(ifg_dct.keys()):
        ifg = ifg_dct.get(hash_id, {})
        ifg_id = ifg.get('_source', {}).get('id', 'MISSING')
        ws.append([ifg_id, hash_id])

def write_hysds_enumerated_date_pairs(wb, acq_list_dct):
    '''writes the sheet that lists all the date pairs from the acquisition lists'''
    ws = wb.create_sheet('HySDS Enumerated Date Pairs')
    ws.append('date pair')
    for id_hash in sort_into_hash_list(acq_list_dct):
        date_pair = gen_date_pair(acq_list_dct.get(id_hash))
        ws.append([date_pair])

def gen_product_met(aoi, product_id, track):
    '''generates the appropriate product json files in the product directory'''
    location = aoi.get('_source', {}).get('location', False)
    starttime = aoi.get('_source', {}).get('starttime', False)
    endtime = aoi.get('_source', {}).get('endtime', False)
    ds_json = {'label': product_id, 'version': VERSION, 'starttime':starttime, 'endtime':endtime, 'location':location}
    outpath = os.path.join(product_id, '{}.dataset.json'.format(product_id))
    with open(outpath, 'w') as outf:
        json.dump(ds_json, outf)
    met_json = {'track_number': track}
    outpath = os.path.join(product_id, '{}.met.json'.format(product_id))
    with open(outpath, 'w') as outf:
        json.dump(met_json, outf)

def validate_enumeration(date_pair_string):
    '''validates the enumeration date pair list to be the appropriate format. Returns as a list sorted by endtime'''
    date_pairs = date_pair_string.replace(' ', '').replace('_', '-').split(',')
    pair_dict = {}
    output_pairs = []
    for date_pair in date_pairs:
        dates = date_pair.split('-')
        if len(dates) < 2:
            print('Failed parsing date pair: {}. skipping.'.format(date_pair))
            continue
        first_date = dateutil.parser.parse(dates[0])
        second_date = dateutil.parser.parse(dates[1])
        if first_date < second_date:
            first_date, second_date = second_date, first_date
        output_date = '{}-{}'.format(first_date.strftime('%Y%m%d'), second_date.strftime('%Y%m%d'))
        pair_dict[first_date] = output_date
    for key in sorted(pair_dict.keys()):
        output_pairs.append(pair_dict.get(key))
    return output_pairs

def sort_date_pair_list(date_pair_list):
    '''sorts a list of date pair strings by the end date'''
    date_dict = {}
    output_list = []
    for date in date_pair_list:
        end_date = re.match('^([0-9]*)-', date).group(1)
        date_dict[end_date] = date
    for key in sorted(date_dict.keys()):
        output_list.append(date_dict.get(key))
    return output_list

def filter_hashes(obj_list, allowed_hashes):
    '''filters out all objects in the object list that aren't storing any of the allowed hashes.'''
    filtered_objs = []
    for obj in obj_list:
        full_id_hash = get_hash(obj)
        if full_id_hash in allowed_hashes:
            filtered_objs.append(obj)
    return filtered_objs

def store_by_hash(obj_list):
    '''returns a dict where the objects are stored by their full_id_hash. drops duplicates.'''
    result_dict = {}
    for obj in obj_list:
        full_id_hash = get_hash(obj)
        if full_id_hash in list(result_dict.keys()):
            result_dict[full_id_hash] = get_most_recent(obj, result_dict.get(full_id_hash))
        else:
            result_dict[full_id_hash] = obj
    return result_dict

def get_most_recent(obj1, obj2):
    '''returns the object with the most recent ingest time'''
    ctime1 = dateutil.parser.parse(obj1.get('_source', {}).get('creation_timestamp', False))
    ctime2 = dateutil.parser.parse(obj2.get('_source', {}).get('creation_timestamp', False))
    if ctime1 > ctime2:
        return obj1
    return obj2

def store_by_id(obj_list):
    '''returns a dict where the objects are stored by their object id'''
    result_dict = {}
    for obj in obj_list:
        obj_id = obj.get('_source', {}).get('id', False)
        if obj_id:
            result_dict[obj_id] = obj
    return result_dict

def sort_by_track(es_result_list):
    '''
    Goes through the objects in the result list, and places them in an dict where key is track
    '''
    sorted_dict = {}
    for result in es_result_list:
        track = get_track(result)
        if track in list(sorted_dict.keys()):
            sorted_dict.get(track, []).append(result)
        else:
            sorted_dict[track] = [result]
    return sorted_dict

def store_by_slc_id(obj_list):
    '''returns a dict where acquisitions are stored by their slc id'''
    result_dict = {}
    for obj in obj_list:
        slc_id = obj.get('_source', {}).get('metadata', {}).get('title', False)
        if slc_id:
            result_dict[slc_id] = obj
    return result_dict

def store_by_gunw(obj_list):
    '''returns a dict where the key is GUNW id and the value is the AOI_TRACK id'''
    result_dict = {}
    for obj in obj_list:
        aoi_track_id = obj.get('_source', {}).get('id', False)
        gunw_ids = obj.get('_source', {}).get('metadata', {}).get('s1-gunw-ids', [])
        for gunw_id in gunw_ids:
            result_dict[gunw_id] = aoi_track_id
    return result_dict

def get_track(es_obj):
    '''returns the track from the elasticsearch object'''
    es_ds = es_obj.get('_source', {})
    #iterate through ds
    track_met_options = ['track_number', 'track', 'trackNumber', 'track_Number']
    for tkey in track_met_options:
        track = es_ds.get(tkey, False)
        if track:
            return track
    #if that doesn't work try metadata
    es_met = es_ds.get('metadata', {})
    for tkey in track_met_options:
        track = es_met.get(tkey, False)
        if track:
            return track
    raise Exception('unable to find track for: {}'.format(es_obj.get('_id', '')))

def store_by_date_pair(obj_list):
    '''returns a dict where the objects are stored by their date_pair'''
    result_dict = {}
    for obj in obj_list:
        date_pair = gen_date_pair(obj)
        result_dict[date_pair] = obj
    return result_dict

def gen_date_pair(obj):
    '''returns the date pair string for the input object'''
    st = dateutil.parser.parse(obj.get('_source').get('starttime')).strftime('%Y%m%d')
    et = dateutil.parser.parse(obj.get('_source').get('endtime')).strftime('%Y%m%d')
    return '{}-{}'.format(et, st)

def sort_into_hash_list(obj_dict):
    '''builds a list of hashes where the hashes are sorted by the objects endtime'''
    sorted_obj = sorted(list(obj_dict.keys()), key=lambda x: get_endtime(obj_dict.get(x)), reverse=True)
    return sorted_obj#[obj.get('_source', {}).get('metadata', {}).get('full_id_hash', '') for obj in sorted_obj]

def get_endtime(obj):
    '''returns the endtime'''
    return dateutil.parser.parse(obj.get('_source', {}).get('endtime'))

def get_hash(es_obj):
    '''retrieves the full_id_hash. if it doesn't exists, it
        attempts to generate one'''
    full_id_hash = es_obj.get('_source', {}).get('metadata', {}).get('full_id_hash', False)
    if full_id_hash:
        return full_id_hash
    return gen_hash(es_obj)

def gen_hash(es_obj):
    '''copy of hash used in the enumerator'''
    met = es_obj.get('_source', {}).get('metadata', {})
    master_slcs = met.get('master_scenes', met.get('reference_scenes', False))
    slave_slcs = met.get('slave_scenes', met.get('secondary_scenes', False))
    master_ids_str = ""
    slave_ids_str = ""
    for slc in sorted(master_slcs):
        if isinstance(slc, tuple) or isinstance(slc, list):
            slc = slc[0]
        if master_ids_str == "":
            master_ids_str = slc
        else:
            master_ids_str += " "+slc
    for slc in sorted(slave_slcs):
        if isinstance(slc, tuple) or isinstance(slc, list):
            slc = slc[0]
        if slave_ids_str == "":
            slave_ids_str = slc
        else:
            slave_ids_str += " "+slc
    id_hash = hashlib.md5(json.dumps([master_ids_str, slave_ids_str]).encode("utf8")).hexdigest()
    return id_hash

def get_objects(object_type, aoi, track_number=False):
    '''returns all objects of the object type ['ifg, acq-list, 'ifg-blacklist'] that intersect both
    temporally and spatially with the aoi'''
    #determine index
    idx = IDX_DCT.get(object_type)
    starttime = aoi.get('_source', {}).get('starttime')
    endtime = aoi.get('_source', {}).get('endtime')
    location = aoi.get('_source', {}).get('location')
    grq_ip = app.conf['GRQ_ES_URL'].replace(':9200', '').replace('http://', 'https://')
    grq_url = '{0}/es/{1}/_search'.format(grq_ip, idx)
    track_field = 'track_number'
    if object_type == 'slc' and track_number:
        track_field = 'trackNumber'
    if track_number:
        grq_query = {"query":{"filtered":{"query":{"geo_shape":{"location": {"shape":location}}},
                     "filter":{"bool":{"must":[{"term":{"metadata.{}".format(track_field):track_number}},
                     {"range":{"endtime":{"gte":starttime}}}, {"range":{"starttime":{"lte":endtime}}}]}}}},
                     "from":0,"size":1000}
    else:
        grq_query = {"query":{"filtered":{"query":{"geo_shape":{"location": {"shape":location}}},
                     "filter":{"bool":{"must":[{"range":{"endtime":{"gte":starttime}}},
                     {"range":{"starttime":{"lte":endtime}}}]}}}},
                     "from":0,"size":1000}
    if object_type == 'audit_trail' or object_type == 'aoi_track':
        grq_query = {"query":{"bool":{"must":[{"term":{"metadata.aoi.raw": aoi.get('_source').get('id')}},{"term":{"metadata.track_number": track_number}}]}},"from":0,"size":1000}
    results = query_es(grq_url, grq_query)
    return results

def query_es(grq_url, es_query):
    '''
    Runs the query through Elasticsearch, iterates until
    all results are generated, & returns the compiled result
    '''
    # make sure the fields from & size are in the es_query
    if 'size' in list(es_query.keys()):
        iterator_size = es_query['size']
    else:
        iterator_size = 10
        es_query['size'] = iterator_size
    if 'from' in list(es_query.keys()):
        from_position = es_query['from']
    else:
        from_position = 0
        es_query['from'] = from_position
    #run the query and iterate until all the results have been returned
    #print('querying: {}\n{}'.format(grq_url, json.dumps(es_query)))
    response = requests.post(grq_url, data=json.dumps(es_query), verify=False)
    #print('status code: {}'.format(response.status_code))
    #print('response text: {}'.format(response.text))
    response.raise_for_status()
    results = json.loads(response.text, encoding='ascii')
    results_list = results.get('hits', {}).get('hits', [])
    total_count = results.get('hits', {}).get('total', 0)
    for i in range(iterator_size, total_count, iterator_size):
        es_query['from'] = i
        #print('querying: {}\n{}'.format(grq_url, json.dumps(es_query)))
        response = requests.post(grq_url, data=json.dumps(es_query), timeout=60, verify=False)
        response.raise_for_status()
        results = json.loads(response.text, encoding='ascii')
        results_list.extend(results.get('hits', {}).get('hits', []))
    return results_list

def get_aoi(aoi_id, aoi_index):
    '''
    retrieves the AOI from ES
    '''
    grq_ip = app.conf['GRQ_ES_URL'].replace(':9200', '').replace('http://', 'https://')
    grq_url = '{0}/es/{1}/_search'.format(grq_ip, aoi_index)
    es_query = {"query":{"bool":{"must":[{"term":{"id.raw":aoi_id}}]}}}
    result = query_es(grq_url, es_query)
    if len(result) < 1:
        raise Exception('Found no results for AOI: {}'.format(aoi_id))
    return result[0]

def load_context():
    '''loads the context file into a dict'''
    try:
        context_file = '_context.json'
        with open(context_file, 'r') as fin:
            context = json.load(fin)
        return context
    except:
        raise Exception('unable to parse _context.json from work directory')


if __name__ == '__main__':
    main()